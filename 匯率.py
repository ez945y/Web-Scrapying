import requests
from bs4 import BeautifulSoup
import openpyxl
from requests.models import default_hooks
def search(dollar):
    url = f'https://www.google.com/search?q={dollar}'
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36'
    web = requests.get(url, headers = {
    'user-agent':user_agent
    })
    soup= BeautifulSoup(web.text,'html5lib')
    ele = soup.find('span', class_='DFlfde SwHCTb')
    if ele:
        return(f"當前台幣兌{dollar}匯率為:{ele.text}元")
    else:
        return None
dollars=[]
workbook = openpyxl.load_workbook("C:\\Users\\admin\\Desktop\\匯率表.xlsx")
sheet = workbook['匯率']
mxR = sheet.max_row
for num in range(2,mxR+1):
    dollars.append(sheet.cell(row = num,column=1).value)
x=2
for dollar in dollars:
    sheet.cell(row=x, column=2).value = search(dollar)
    x += 1
workbook.save("C:\\Users\\admin\\Desktop\\匯率表.xlsx")

